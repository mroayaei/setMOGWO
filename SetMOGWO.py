import multiprocessing as mprc
import ndlib.models.ModelConfig as mc
from IM import *
from Domination import *
from Hypercube import *
from Dataset.dataset import *
from solution import solution
import math
from collections import Counter
from heapq import nlargest
import time
import openpyxl as xl
import fitness as ft
import FairnessUtil as fu

__program__ = None

class Wolfs:
    def __init__(self, greywolf, instance, attribute):
        self.greyWolf = greywolf
        self.instance = instance
        self.attribute = attribute


def CalculateWolfFitness_1(greywolfs):                  
        return multi_fobj_1(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_2(greywolfs):                  
        return multi_fobj_2_ir(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_3(greywolfs):                  
        return multi_fobj_2_im(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_4(greywolfs):                  
        return multi_fobj_2_is(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_5(greywolfs):          
        return multi_fobj_3_imr(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_6(greywolfs):          
        return multi_fobj_3_ims(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_7(greywolfs):              
        return multi_fobj_3_irs(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)    
def CalculateWolfFitness_8(greywolfs):              
        return multi_fobj_4(list(greywolfs.greyWolf.Position), greywolfs.instance, greywolfs.attribute)

if __name__ == '__main__':         
    attributes = ['age','ethnicity','gender']    
    progs = [2,3,4,5,6,7,8]
    for idx ,program in enumerate(progs):                
        for _,attribute in enumerate(attributes):                                      
            print("program " +str(program)+ " starting for attribute: " + attribute)
            book = xl.Workbook()
            wbs = book.create_sheet(attribute)
            wbs.cell(column=1, row=1).value = 'ActivePercent'
            wbs.cell(column=2, row=1).value = 'MaximinPercent'
            wbs.cell(column=3, row=1).value = 'TimePercent'
            wbs.cell(column=4, row=1).value = 'RationalityPercent'
            wbs.cell(column=5, row=1).value = 'RationalityViolation'
            wbs.cell(column=6, row=1).value = 'PickleNo'       
            curExcelRow = 1
            numOfPcikles = 24
            for p in range(0, numOfPcikles):              
                print("pickle " +str(p)+" starting...")
                pickleNo = p            
                instance = prepare_dataset_SAV(pickleNo)
                GreyWolves_num = 500
                MaxIt = 100
                dim = instance.budget
                Alpha_score = float("-inf")
                Beta_score = float("-inf")
                Delta_score = float("-inf")


                Alpha_pos = set()
                Beta_pos = set()
                Delta_pos = set()

                graph = instance.graph
                config = mc.Configuration()
                removed_edges = []
                for e in graph.edges():
                    # we set the weight of every incoming edge of a node v to be equal to 1/dv,
                    # where dv is the in-degree of node v.
                    if(random.random() <= instance.edgeThreshold):
                        config.add_edge_configuration("threshold", e, 1)
                    else:
                        removed_edges.append(e)
                instance.config = config
                graph.remove_edges_from(removed_edges)
                graph.remove_edges_from(nx.selfloop_edges(graph))

                    
                timerStart = time.time()         

                deg_centrality = nx.degree_centrality(graph)
                #eigen_centrality = nx.eigenvector_centrality(graph)
                eigen_centrality = nx.eigenvector_centrality_numpy(graph,max_iter=100)
                Cdict = dict(Counter(deg_centrality) + Counter(eigen_centrality))
                #Cdict = dict(sorted(Cdict.items(), key=lambda item: item[1], reverse=True))
                largest = nlargest(instance.budget, Cdict, key=Cdict.get)
                allWolfs = []

                # Repository Size
                Archive_size = 100
                # Grid Inflation Parameter
                alpha = 0.1
                # Number of Grids per each Dimension
                nGrid = 10
                # Leader Selection Pressure Parameter
                beta = 4
                # Extra (to be deleted) Repository Member Selection Pressure
                gamma = 2

                # Initialization
                #all items
                seq = list(instance.graph.nodes() - largest)
                seq_set = [set(seq)] * GreyWolves_num
                largest_number = math.ceil(0 * instance.budget)     
                # initialize wolfs
                for _ in range(GreyWolves_num):
                    wolf = GreyWolf()
                    wolf.Position = set(random.sample(largest, k=largest_number) + random.sample(seq, k=instance.budget - largest_number))
                    wolf.Best_Position = wolf.Position
                    wolf.Velocity = 0
                    allWolfs.append(Wolfs(wolf, instance, attribute))

                # initialize w.r.t the number of iteration
                # if cpu count < 10 use pool else it is better to use processes
                #pool = mprc.Pool(processes=1)
                pool = mprc.Pool(processes=mprc.cpu_count())
                # res is a list that every its item is a fitness value and its corresponding grey wolf
                if program == 1:
                        res= list(pool.imap(CalculateWolfFitness_1, allWolfs))
                elif program == 2:
                        res= list(pool.imap(CalculateWolfFitness_2, allWolfs))
                elif program == 3:
                    res= list(pool.imap(CalculateWolfFitness_3, allWolfs))
                elif program == 4:
                    res= list(pool.imap(CalculateWolfFitness_4, allWolfs))
                elif program == 5:
                    res= list(pool.imap(CalculateWolfFitness_5, allWolfs))
                elif program == 6:
                    res= list(pool.imap(CalculateWolfFitness_6, allWolfs))
                elif program == 7:
                    res= list(pool.imap(CalculateWolfFitness_7, allWolfs))
                elif program == 8:
                    res= list(pool.imap(CalculateWolfFitness_8, allWolfs))
                
                pool.close()
                pool.join()
                i = 0
                GreyWolves = []
                for obj in allWolfs:
                    obj.greyWolf.Cost = res[i][:-1]
                    obj.greyWolf.Best_Cost = res[i][:-1]
                    obj.greyWolf.RationalityViolation = res[i][-1]
                    GreyWolves.append(obj.greyWolf)
                    i += 1

                GreyWolves = DetermineDomination(GreyWolves)

                Archive = GetNonDominatedParticles(GreyWolves)

                # segmentation of archive
                Archive_costs = GetCosts(Archive)
                G = CreateHypercubes(Archive_costs, nGrid, alpha)

                for i in range(len(Archive)):
                    Archive[i].GridIndex, Archive[i].GridSubIndex = GetGridIndex(Archive[i], G)

                # clear rep2
                rep2 = []
                # clear rep3
                rep3 = []

                # Choose the alpha, beta, and delta grey wolf
                Delta = SelectLeader(Archive, beta)
                Beta = SelectLeader(Archive, beta)
                Alpha = SelectLeader(Archive, beta)

                """
                If there are less than three solutions in the least crowded
                hypercube, the second least crowded hypercube is also found
                to choose other leaders from.
                """
                if len(Archive) > 1:
                    for newi in range(len(Archive)):
                        if np.sum(Delta.Position != Archive[newi].Position) != 0:
                            rep2.append(Archive[newi])
                    Beta = SelectLeader(rep2, beta)

                """
                This scenario is the same if the second least crowded hypercube
                has one solution, so the delta leader should be chosen from the
                third least crowded hypercube.
                """
                if len(Archive) > 2:
                    for newi in range(len(rep2)):
                        if np.sum(Beta.Position != rep2[newi].Position) != 0:
                            rep3.append(rep2[newi])
                    Alpha = SelectLeader(rep3, beta)

                # MOGWO main loop
                Alpha_pos = Alpha.Position.copy()
                Beta_pos = Beta.Position.copy()
                Delta_pos = Delta.Position.copy()

                for it in range(MaxIt):
                    print("iteration " +str(it)+" starting...")
                    iterstarttime = time.time()

                    print('In iteration ' + str(it) + ' : Number of solutions in the archive = ' + str(len(Archive)))
                    for i in range(len(Archive)):
                        print(Archive[i].Cost)
                    
                        if Archive[i].Cost[3] == 0 and Archive[i].RationalityViolation > 0:
                            print('Rationality Violation : ' + str(Archive[i].RationalityViolation))
                    

                    a = (MaxIt-it)*(2 / MaxIt)
                    for i in range(GreyWolves_num):
                        if len(seq_set[i]) < 5 * dim:
                            # refill set_seq
                            seq_set[i] = set(seq)

                        # generate random numbers
                        r1 = random.random()
                        r2 = random.random()
                        C = 2 * r2
                        A = 2 * a * r1 - a

                        # determine which leader is closer
                        intersec_alpha = len(Alpha_pos.intersection(GreyWolves[i].Position))
                        intersec_beta = len(Beta_pos.intersection(GreyWolves[i].Position))
                        intersec_delta = len(Delta_pos.intersection(GreyWolves[i].Position))
                        
                        leader = set()
                        if intersec_alpha > intersec_beta and intersec_alpha > intersec_delta:
                            leader = Alpha_pos
                        elif (intersec_beta > intersec_alpha and intersec_beta > intersec_delta):
                            leader = Beta_pos
                        elif (intersec_delta > intersec_alpha and intersec_delta > intersec_beta):
                            leader = Delta_pos
                        elif (intersec_alpha == intersec_beta == intersec_delta):
                            leader = Alpha_pos.union(Beta_pos.union(Delta_pos))
                        elif (intersec_alpha == intersec_beta):
                            leader = Alpha_pos.union(Beta_pos)
                        elif (intersec_alpha == intersec_delta):
                            leader = Alpha_pos.union(Delta_pos)
                        elif (intersec_beta == intersec_delta):
                            leader = Beta_pos.union(Delta_pos)
                        else:
                            leader = Alpha_pos.union(Beta_pos.union(Delta_pos))                     

                        leader_items = list(leader)

                        CBound = math.ceil(C * len(leader_items))
                        CBound = min(CBound, len(leader_items))
                        #CBound = len(leader_items)
                        leader_set = set(random.sample(leader, k=CBound))

                        if (abs(A) < 1):
                            A = abs(A)
                            # exploitation : get closer to leaders
                            # select new items which are not in omega
                            new_items = leader_set - GreyWolves[i].Position
                            old_items = GreyWolves[i].Position - leader_set
                            D = len(old_items)
                            selecteds_num = abs(math.ceil(A * D))
                            selecteds_num = min(selecteds_num, len(old_items), len(new_items))
                            GreyWolves[i].Position.difference_update(random.sample(old_items, k=selecteds_num))
                            GreyWolves[i].Position = GreyWolves[i].Position.union(random.sample(new_items, k=selecteds_num))
                        else:
                            # exploration: get farther from leaders
                            # select new items randomly which are not currentlly in Greywolves[i]
                            new_items = seq_set[i] - (GreyWolves[i].Position.union(leader_set))
                            old_items = leader_set.intersection(GreyWolves[i].Position)
                            # D = int(len(new_items)/3)
                            D = len(GreyWolves[i].Position - leader_set)
                            selecteds_num = math.ceil(abs(A * D))
                            selecteds_num = min(selecteds_num, len(old_items), len(new_items))
                            if (selecteds_num != 0):
                                GreyWolves[i].Position.difference_update(random.sample(old_items, k=selecteds_num))
                                news = random.sample(new_items, k=selecteds_num)
                                GreyWolves[i].Position = GreyWolves[i].Position.union(news)
                                seq_set[i] -= set(news)                            
                                            
                    #pool = mprc.Pool(processes=1)
                    pool = mprc.Pool(processes=mprc.cpu_count())
                    # res is a list that every its item is a fitness value and its corresponding grey wolf
                    if program == 1:
                        res= list(pool.imap(CalculateWolfFitness_1, allWolfs))
                    elif program == 2:
                        res= list(pool.imap(CalculateWolfFitness_2, allWolfs))
                    elif program == 3:
                        res= list(pool.imap(CalculateWolfFitness_3, allWolfs))
                    elif program == 4:
                        res= list(pool.imap(CalculateWolfFitness_4, allWolfs))
                    elif program == 5:
                        res= list(pool.imap(CalculateWolfFitness_5, allWolfs))
                    elif program == 6:
                        res= list(pool.imap(CalculateWolfFitness_6, allWolfs))
                    elif program == 7:
                        res= list(pool.imap(CalculateWolfFitness_7, allWolfs))
                    elif program == 8:
                        res= list(pool.imap(CalculateWolfFitness_8, allWolfs))
                    pool.close()
                    pool.join()

                    i = 0
                    GreyWolves = []
                    for obj in allWolfs:
                        obj.greyWolf.Cost = res[i][:-1]
                        obj.greyWolf.Best_Cost = res[i][:-1]
                        obj.greyWolf.RationalityViolation = res[i][-1]
                        GreyWolves.append(obj.greyWolf)
                        i += 1

                    GreyWolves = DetermineDomination(GreyWolves)

                    non_dominated_wolves = GetNonDominatedParticles(GreyWolves)

                    # add new non-dominated wolves to the archive
                    Archive = np.concatenate((Archive, non_dominated_wolves))

                    Archive = DetermineDomination(Archive)
                    Archive = GetNonDominatedParticles(Archive)

                    uniq = []
                    uniqArchive = []
                    for item in Archive:
                        if list(item.Cost) not in uniq:
                            uniqArchive.append(item)
                            uniq.append(list(item.Cost))

                    Archive = uniqArchive
                    #
                    for i in range(len(Archive)):
                        Archive[i].GridIndex, Archive[i].GridSubIndex = GetGridIndex(Archive[i], G)

                    if len(Archive) > Archive_size:
                        EXTRA = len(Archive) - Archive_size
                        Archive = DeleteFromRep(Archive, EXTRA, gamma)

                        Archive_costs = GetCosts(Archive)
                        G = CreateHypercubes(Archive_costs, nGrid, alpha)

                    # clear rep2
                    rep2 = []
                    # clear rep3
                    rep3 = []

                    # Choose the alpha, beta, and delta grey wolf
                    Delta = SelectLeader(Archive, beta)
                    Beta = SelectLeader(Archive, beta)
                    Alpha = SelectLeader(Archive, beta)

                    """
                    If there are less than three solutions in the least crowded
                    hypercube, the second least crowded hypercube is also found
                    to choose other leaders from.
                    """
                    if len(Archive) > 1:
                        for newi in range(len(Archive)):
                            if np.sum(Delta.Position != Archive[newi].Position) != 0:
                                rep2.append(Archive[newi])
                        Beta = SelectLeader(rep2, beta)

                    """
                    This scenario is the same if the second least crowded hypercube
                    has one solution, so the delta leader should be chosen from the
                    third least crowded hypercube.
                    """
                    if len(Archive) > 2:
                        for newi in range(len(rep2)):
                            if np.sum(Beta.Position != rep2[newi].Position) != 0:
                                rep3.append(rep2[newi])
                        Alpha = SelectLeader(rep3, beta)

                    Alpha_pos = Alpha.Position.copy()
                    Beta_pos = Beta.Position.copy()
                    Delta_pos = Delta.Position.copy()

                    # save results

                    # Results

                    costs = GetCosts(GreyWolves)
                    Archive_costs = GetCosts(Archive)
                
                    print('Iteration Time : ' + str(time.time() - iterstarttime))
                #print("best solution is: " + str(Archive[0].Position))
                timerEnd = time.time()        


                # for each pickle
                for i in range(len(Archive)):
                    print('before')
                    Archive[i].Cost = multi_fobj_4(Archive[i].Position,instance,attribute)
                    print('after')
                    wbs.cell(column=1, row=curExcelRow + 1 + i).value = str(Archive[i].Cost[0])
                    wbs.cell(column=2, row=curExcelRow + 1 + i).value = str(Archive[i].Cost[1])
                    wbs.cell(column=3, row=curExcelRow + 1 + i).value = str(Archive[i].Cost[2])
                    wbs.cell(column=4, row=curExcelRow + 1 + i).value = str(Archive[i].Cost[3])
                    wbs.cell(column=5, row=curExcelRow + 1 + i).value = str(Archive[i].Cost[4])
                    wbs.cell(column=6, row=curExcelRow + 1 + i).value = str(pickleNo)                
                curExcelRow = len(Archive) + curExcelRow
                book.save(attribute +"_"+str(program)+ '.xlsx')   
     




